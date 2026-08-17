"""Load the scraped JSONL archive into the local Postgres database.

Idempotent: truncates and reloads every table from downloads/ + data/.
Usage: .venv/bin/python db/load.py
"""

import json
import os
from datetime import date, datetime
from pathlib import Path

import psycopg
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_URL = "postgres://bes:bes@localhost:5435/bes"

DT_FULL = "%d %b %Y %H:%M:%S"   # 03 Aug 2026 09:14:48
DT_MIN = "%d %b %Y %H:%M"       # 31 May 2022 02:00 (register reports)
DT_DATE = "%d %b %Y"            # 05 Jan 2026


def read_jsonl(path: Path) -> list[dict]:
    """Return all JSON objects from a JSONL file, skipping blank lines."""
    with path.open() as f:
        return [json.loads(line) for line in f if line.strip()]


def parse_ts(value: str | None) -> datetime | None:
    """Parse a BES datetime string; return None when missing or unparseable."""
    if not value:
        return None
    for fmt in (DT_FULL, DT_MIN, DT_DATE):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value)  # e.g. 2021-12-31 in daily stats
    except ValueError:
        return None


def parse_date(value: str | None) -> date | None:
    """Parse a BES date string into a datetime.date; None when unparseable."""
    ts = parse_ts(value)
    return ts.date() if ts else None


def parse_bool(value: str | None) -> bool | None:
    """Map BES 'Y'/'N' flags to booleans; None when absent."""
    if value == "Y":
        return True
    if value == "N":
        return False
    return None


def to_int(value) -> int | None:
    """Cast to int, tolerating strings/empties; None when not numeric."""
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def load_permits(cur) -> int:
    """Load data/ptw.jsonl into permits."""
    rows = read_jsonl(ROOT / "data" / "ptw.jsonl")
    cur.executemany(
        "INSERT INTO permits VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
        [
            (
                int(r["apply_id"]),
                r["serial_no"],
                r["company"],
                r.get("title"),
                r["type"],
                r.get("status"),
                parse_ts(r.get("date_of_application")),
                r.get("pdf_file"),
            )
            for r in rows
        ],
    )
    return len(rows)


def load_approval_trail(cur) -> tuple[int, int, int]:
    """Load downloads/ptw/extras.jsonl into approval_steps + evidence links."""
    steps = checklists = attachments = 0
    for r in read_jsonl(ROOT / "downloads" / "ptw" / "extras.jsonl"):
        apply_id = int(r["apply_id"])
        cur.executemany(
            "INSERT INTO approval_steps VALUES (%s,%s,%s,%s)",
            [(apply_id, int(s["step"]), s["role"], s["approver_name"]) for s in r["workflow"]],
        )
        steps += len(r["workflow"])
        chk = [(apply_id, i) for i in set(r.get("checklist_report_ids") or [])]
        cur.executemany("INSERT INTO permit_checklists VALUES (%s,%s) ON CONFLICT DO NOTHING", chk)
        checklists += len(chk)
        att = [(apply_id, i) for i in set(r.get("attachment_ids") or [])]
        cur.executemany("INSERT INTO permit_attachments VALUES (%s,%s) ON CONFLICT DO NOTHING", att)
        attachments += len(att)
    return steps, checklists, attachments


def load_staff(cur) -> int:
    """Load downloads/staff/staff.jsonl into staff."""
    rows = read_jsonl(ROOT / "downloads" / "staff" / "staff.jsonl")
    cur.executemany(
        "INSERT INTO staff VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        [
            (
                int(r["staff_id"]),
                r["full_name"],
                r.get("badge_no") or None,
                r.get("mobile_no") or None,
                r.get("nric_fin") or None,
                r.get("id_type"),
                r.get("nationality"),
                r.get("designation"),
                parse_bool(r.get("secondment")),
                parse_date(r.get("created_on")),
            )
            for r in rows
        ],
    )
    return len(rows)


def load_equipment(cur) -> int:
    """Load downloads/equipment/equipment.jsonl into equipment."""
    rows = read_jsonl(ROOT / "downloads" / "equipment" / "equipment.jsonl")
    cur.executemany(
        "INSERT INTO equipment VALUES (%s,%s,%s,%s,%s)",
        [
            (
                int(r["equipment_id"]),
                r.get("equipment_type"),
                r.get("registration_no"),
                r.get("equipment_name"),
                parse_date(r.get("created_on")),
            )
            for r in rows
        ],
    )
    return len(rows)


def load_company_documents(cur) -> int:
    """Load downloads/company/company.jsonl into company_documents."""
    rows = read_jsonl(ROOT / "downloads" / "company" / "company.jsonl")
    cur.executemany(
        "INSERT INTO company_documents VALUES (%s,%s,%s)",
        [
            (int(r["document_id"]), r["document_name"], parse_ts(r.get("uploaded_on")))
            for r in rows
        ],
    )
    return len(rows)


def load_monthly_stats(cur) -> int:
    """Load downloads/statistics/statistics.jsonl into monthly_stats (frozen snapshot)."""
    rows = read_jsonl(ROOT / "downloads" / "statistics" / "statistics.jsonl")
    cur.executemany(
        "INSERT INTO monthly_stats VALUES (%s,%s,%s,%s)",
        [
            (
                r["window"],
                parse_date(r.get("start_date")),
                parse_date(r.get("end_date")),
                json.dumps(r["counters"]),
            )
            for r in rows
        ],
    )
    return len(rows)


def load_company_daily_stats(cur) -> int:
    """Load downloads/statistics/company_daily.jsonl into company_daily_stats."""
    rows = read_jsonl(ROOT / "downloads" / "statistics" / "company_daily.jsonl")
    cur.executemany(
        "INSERT INTO company_daily_stats VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        [
            (
                parse_date(r["date"]),
                r["company"],
                to_int(r["ptw_cnt"]),
                to_int(r["ptw_staff"]),
                to_int(r["tbm_cnt"]),
                r["tbm_staff"],
                to_int(r["inspection_cnt"]),
                to_int(r["meeting_cnt"]),
                to_int(r["meeting_staff"]),
                to_int(r["training_cnt"]),
                to_int(r["training_staff"]),
                to_int(r["ra_cnt"]),
                to_int(r["checklist_cnt"]),
                to_int(r["incident_cnt"]),
            )
            for r in rows
        ],
    )
    return len(rows)


def load_monthly_permit_status(cur) -> int:
    """Aggregate downloads/ptwreports/*.xlsx registers into monthly_permit_status.

    Register layout: title row, two header rows, data from row 4;
    status lives in column index 8. Rows without an S/N or status are skipped.
    """
    from openpyxl import load_workbook

    rows: dict[tuple[str, str], int] = {}
    for path in sorted((ROOT / "downloads" / "ptwreports").glob("*.xlsx")):
        month = path.stem
        wb = load_workbook(path, read_only=True)
        for row in wb.active.iter_rows(min_row=4, values_only=True):
            if row[0] is None or not row[8]:
                continue
            key = (month, str(row[8]).strip())
            rows[key] = rows.get(key, 0) + 1
        wb.close()
    cur.executemany(
        "INSERT INTO monthly_permit_status VALUES (%s,%s,%s)",
        [(m, s, n) for (m, s), n in sorted(rows.items())],
    )
    return len(rows)


def load_permit_register(cur) -> int:
    """Load full register rows from downloads/ptwreports/*.xlsx into permit_register.

    Register layout: title row, two header rows, data from row 4. Columns:
    S/N, location, work type, company, serial no, start, end, approver, status.
    Rows without a serial number or status, and duplicate serials within a
    month, are skipped.
    """
    from openpyxl import load_workbook

    def cell(value) -> str | None:
        s = str(value).strip() if value is not None else ""
        return s or None

    seen: set[tuple[str, str]] = set()
    batch = []
    for path in sorted((ROOT / "downloads" / "ptwreports").glob("*.xlsx")):
        month = path.stem
        wb = load_workbook(path, read_only=True)
        for row in wb.active.iter_rows(min_row=4, values_only=True):
            serial, status = cell(row[4]), cell(row[8])
            if not serial or not status or (month, serial) in seen:
                continue
            seen.add((month, serial))
            batch.append((
                month, serial, cell(row[1]), cell(row[2]), cell(row[3]),
                parse_ts(cell(row[5])), parse_ts(cell(row[6])), cell(row[7]), status,
            ))
        wb.close()
    cur.executemany(
        "INSERT INTO permit_register VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)", batch
    )
    return len(batch)


def load_staff_documents(cur) -> int:
    """Load per-worker document attachments from downloads/staff/staff_docs.jsonl.

    One JSONL line per worker ({"staff_id": ..., "documents": [...]}); rows are
    flattened into staff_documents. Duplicates within the file are skipped.
    """
    path = ROOT / "downloads" / "staff" / "staff_docs.jsonl"
    if not path.exists():
        return 0
    seen: set[tuple] = set()
    batch = []
    for line in path.read_text().splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        for d in rec["documents"]:
            row = (
                rec["staff_id"], d.get("document_no"), d.get("doc_type"),
                parse_date(d.get("issue_date")), parse_date(d.get("expiry_date")),
                parse_date(d.get("uploaded_on")),
            )
            if row in seen:
                continue
            seen.add(row)
            batch.append(row)
    cur.executemany(
        "INSERT INTO staff_documents"
        " (staff_id, document_no, doc_type, issue_date, expiry_date, uploaded_on)"
        " VALUES (%s,%s,%s,%s,%s,%s)",
        batch,
    )
    return len(batch)


def load_permit_members(cur) -> int:
    """Load PTW member rows from downloads/ptw/members.jsonl into permit_members.

    One JSONL line per permit ({"apply_id": ..., "members": [...]}); rows are
    flattened. Duplicates within the file are skipped.
    """
    path = ROOT / "downloads" / "ptw" / "members.jsonl"
    if not path.exists():
        return 0
    seen: set[tuple] = set()
    batch = []
    for line in path.read_text().splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        for m in rec["members"]:
            row = (
                rec["apply_id"], m.get("id_number"), m.get("name"),
                m.get("employee_id"), m.get("designation"),
            )
            if row in seen:
                continue
            seen.add(row)
            batch.append(row)
    cur.executemany(
        "INSERT INTO permit_members"
        " (apply_id, id_number, name, employee_id, designation)"
        " VALUES (%s,%s,%s,%s,%s)",
        batch,
    )
    return len(batch)


def main() -> None:
    """Apply schema, truncate, reload every table, and print loaded counts."""
    load_dotenv(ROOT / ".env")
    url = os.environ.get("DATABASE_URL", DEFAULT_URL)
    with psycopg.connect(url) as conn:
        with conn.cursor() as cur:
            cur.execute((ROOT / "db" / "schema.sql").read_text())
            cur.execute(
                "TRUNCATE permits, staff, equipment, company_documents,"
                " monthly_stats, company_daily_stats, monthly_permit_status,"
                " permit_register, staff_documents, permit_members CASCADE"
            )
            counts = {
                "permits": load_permits(cur),
                "staff": load_staff(cur),
                "equipment": load_equipment(cur),
                "company_documents": load_company_documents(cur),
                "monthly_stats": load_monthly_stats(cur),
                "company_daily_stats": load_company_daily_stats(cur),
                "monthly_permit_status": load_monthly_permit_status(cur),
                "permit_register": load_permit_register(cur),
                "staff_documents": load_staff_documents(cur),
                "permit_members": load_permit_members(cur),
            }
            counts["approval_steps"], counts["permit_checklists"], counts["permit_attachments"] = (
                load_approval_trail(cur)
            )
        conn.commit()
    for table, n in counts.items():
        print(f"{table:22} {n:>7}")


if __name__ == "__main__":
    main()
